import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Student Template"

# Read CSV
with open('sample-student-template.csv', 'r', encoding='utf-8') as f:
    reader = csv.reader(f)
    for row in reader:
        ws.append(row)

# Style header row
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column_letter].width = adjusted_width

# Save
wb.save('sample-student-template.xlsx')
print("✓ Created sample-student-template.xlsx")
